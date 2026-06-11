import io
from calendar import month_name
from datetime import date, timedelta
from flask import Blueprint, request, jsonify, send_file
from sqlalchemy import text

from middleware.auth_middleware import admin_required
from config.database import db
from utils.logger import get_logger

analytics_bp = Blueprint("analytics", __name__)
logger       = get_logger(__name__)


# ── Dashboard Summary ──────────────────────────────────────────

@analytics_bp.route("/summary", methods=["GET"])
@admin_required
def summary():
    """GET /api/v1/analytics/summary — today's attendance summary"""
    today = date.today()

    try:
        # Total active employees
        total = db.session.execute(text("""
            SELECT COUNT(*) FROM employees
            WHERE employment_status = 'active'
        """)).scalar() or 0

        # Today's stats.
        # Count "late" from the is_late flag (not the status column) so the
        # dashboard always agrees with the attendance list, even if a row's
        # status text drifted (e.g. manual clock-in edge cases).
        stats = db.session.execute(text("""
            SELECT
                COUNT(*) FILTER (WHERE clock_in IS NOT NULL
                                   AND is_late IS NOT TRUE
                                   AND status != 'on_leave')  AS present,
                COUNT(*) FILTER (WHERE is_late IS TRUE)       AS late,
                COUNT(*) FILTER (WHERE status = 'absent')     AS absent,
                COUNT(*) FILTER (WHERE status = 'on_leave')   AS on_leave
            FROM attendance_logs
            WHERE attendance_date = :today
        """), {"today": today}).fetchone()

        present  = stats.present  if stats else 0
        late     = stats.late     if stats else 0
        absent   = stats.absent   if stats else 0
        on_leave = stats.on_leave if stats else 0

        # Employees not yet clocked in = total - (present + late + on_leave)
        not_clocked = max(0, total - present - late - on_leave)

        attendance_rate = round((present + late) / total * 100, 1) if total > 0 else 0

        return jsonify({
            "total_employees":  total,
            "present":          present,
            "late":             late,
            "absent":           not_clocked,
            "on_leave":         on_leave,
            "attendance_rate":  attendance_rate,
            "date":             today.isoformat(),
        }), 200

    except Exception as e:
        logger.error(f"Summary error: {e}")
        return jsonify({
            "total_employees": 0,
            "present":         0,
            "late":            0,
            "absent":          0,
            "on_leave":        0,
            "attendance_rate": 0,
            "date":            today.isoformat(),
        }), 200


# ── 30-day Attendance Trend ────────────────────────────────────

@analytics_bp.route("/trend", methods=["GET"])
@admin_required
def trend():
    """GET /api/v1/analytics/trend — daily present/absent for last 30 days"""
    days = int(request.args.get("days", 30))
    end_date   = date.today()
    start_date = end_date - timedelta(days=days - 1)

    try:
        rows = db.session.execute(text("""
            SELECT
                attendance_date,
                COUNT(*) FILTER (WHERE status IN ('present','late')) AS present,
                COUNT(*) FILTER (WHERE status = 'absent')            AS absent,
                COUNT(*) FILTER (WHERE status = 'late')              AS late,
                COUNT(*) FILTER (WHERE status = 'on_leave')          AS on_leave
            FROM attendance_logs
            WHERE attendance_date BETWEEN :start AND :end
            GROUP BY attendance_date
            ORDER BY attendance_date ASC
        """), {"start": start_date, "end": end_date}).fetchall()

        # Build a map so we can fill in missing days with zeros
        # Total active employees (for rate calculation)
        total_employees = db.session.execute(text(
            "SELECT COUNT(*) FROM employees WHERE employment_status = 'active'"
        )).scalar() or 1  # avoid div-by-zero

        data_map = {}
        for row in rows:
            d = dict(row._mapping)
            present = int(d["present"] or 0)
            absent  = int(d["absent"]  or 0)
            late    = int(d["late"]    or 0)
            rate    = round((present) / total_employees * 100, 1)
            data_map[str(d["attendance_date"])] = {
                "date":            str(d["attendance_date"]),
                "present":         present,
                "absent":          absent,
                "late":            late,
                "on_leave":        int(d["on_leave"] or 0),
                "attendance_rate": rate,
            }

        # Fill missing days with zeros
        trend_data = []
        current = start_date
        while current <= end_date:
            key = str(current)
            trend_data.append(data_map.get(key, {
                "date":            key,
                "present":         0,
                "absent":          0,
                "late":            0,
                "on_leave":        0,
                "attendance_rate": 0,
            }))
            current += timedelta(days=1)

        return jsonify({
            "trend":      trend_data,
            "start_date": str(start_date),
            "end_date":   str(end_date),
            "days":       days,
        }), 200

    except Exception as e:
        logger.error(f"Trend error: {e}")
        return jsonify({
            "trend":      [],
            "start_date": str(start_date),
            "end_date":   str(end_date),
            "days":       days,
        }), 200


# ── Department Breakdown ───────────────────────────────────────

@analytics_bp.route("/departments", methods=["GET"])
@admin_required
def department_stats():
    """GET /api/v1/analytics/departments"""
    today = date.today()

    try:
        rows = db.session.execute(text("""
            SELECT
                d.name AS department,
                COUNT(DISTINCT e.id)                                              AS total,
                COUNT(al.id) FILTER (WHERE al.status IN ('present','late'))       AS present,
                COUNT(al.id) FILTER (WHERE al.status = 'late')                    AS late,
                COUNT(al.id) FILTER (WHERE al.status = 'absent')                  AS absent
            FROM departments d
            JOIN employees e ON e.department_id = d.id
                AND e.employment_status = 'active'
            LEFT JOIN attendance_logs al ON al.employee_id = e.id
                AND al.attendance_date = :today
            WHERE d.is_active = TRUE
            GROUP BY d.name
            ORDER BY d.name
        """), {"today": today}).fetchall()

        departments = []
        for row in rows:
            d = dict(row._mapping)
            total = int(d["total"] or 0)
            present = int(d["present"] or 0)
            departments.append({
                "department":      d["department"],
                "total":           total,
                "present":         present,
                "late":            int(d["late"]   or 0),
                "absent":          int(d["absent"] or 0),
                "attendance_rate": round(present / total * 100, 1) if total > 0 else 0,
            })

        return jsonify({"departments": departments}), 200

    except Exception as e:
        logger.error(f"Department stats error: {e}")
        return jsonify({"departments": []}), 200


# ── Late Arrivals Report ───────────────────────────────────────

@analytics_bp.route("/late-report", methods=["GET"])
@admin_required
def late_report():
    """GET /api/v1/analytics/late-report?days=30"""
    days       = int(request.args.get("days", 30))
    end_date   = date.today()
    start_date = end_date - timedelta(days=days - 1)

    try:
        rows = db.session.execute(text("""
            SELECT
                e.employee_id,
                e.first_name || ' ' || e.last_name AS name,
                d.name AS department,
                COUNT(*) AS late_count,
                AVG(al.late_minutes) AS avg_late_minutes
            FROM attendance_logs al
            JOIN employees e  ON al.employee_id = e.id
            LEFT JOIN departments d ON e.department_id = d.id
            WHERE al.status = 'late'
              AND al.attendance_date BETWEEN :start AND :end
            GROUP BY e.employee_id, e.first_name, e.last_name, d.name
            ORDER BY late_count DESC
            LIMIT 20
        """), {"start": start_date, "end": end_date}).fetchall()

        report = []
        for row in rows:
            d = dict(row._mapping)
            report.append({
                "employee_id":      d["employee_id"],
                "name":             d["name"],
                "department":       d["department"],
                "late_count":       int(d["late_count"] or 0),
                "avg_late_minutes": round(float(d["avg_late_minutes"] or 0), 1),
            })

        return jsonify({
            "report":     report,
            "start_date": str(start_date),
            "end_date":   str(end_date),
            "days":       days,
        }), 200

    except Exception as e:
        logger.error(f"Late report error: {e}")
        return jsonify({"report": [], "days": days}), 200


# ── Monthly Late Summary (all staff) ───────────────────────────

def _parse_year_month():
    """Read ?year=&month= from query string (defaults to current month)."""
    today = date.today()
    year  = int(request.args.get("year",  today.year))
    month = int(request.args.get("month", today.month))
    if not 1 <= month <= 12:
        raise ValueError
    return year, month


def _month_range(year: int, month: int):
    start = date(year, month, 1)
    end = (date(year + 1, 1, 1) if month == 12
           else date(year, month + 1, 1)) - timedelta(days=1)
    return start, end


def _get_late_summary(year: int, month: int) -> list:
    """All active staff with late counts for the month (zero lates included)."""
    month_start, month_end = _month_range(year, month)
    rows = db.session.execute(text("""
            SELECT
                e.employee_id,
                e.first_name || ' ' || e.last_name AS name,
                d.name AS department,
                COUNT(al.id) FILTER (WHERE al.is_late = TRUE)      AS late_count,
                COALESCE(SUM(al.late_minutes), 0)                  AS total_late_minutes,
                COUNT(al.id) FILTER (WHERE al.clock_in IS NOT NULL) AS days_present
            FROM employees e
            LEFT JOIN departments d ON e.department_id = d.id
            LEFT JOIN attendance_logs al ON al.employee_id = e.id
                AND al.attendance_date BETWEEN :start AND :end
            WHERE e.employment_status = 'active'
            GROUP BY e.employee_id, e.first_name, e.last_name, d.name
            ORDER BY late_count DESC, name ASC
    """), {"start": month_start, "end": month_end}).fetchall()

    summary = []
    for row in rows:
        d = dict(row._mapping)
        summary.append({
            "employee_id":        d["employee_id"],
            "name":               d["name"],
            "department":         d["department"],
            "late_count":         int(d["late_count"] or 0),
            "total_late_minutes": int(d["total_late_minutes"] or 0),
            "days_present":       int(d["days_present"] or 0),
        })
    return summary


@analytics_bp.route("/late-summary", methods=["GET"])
@admin_required
def late_summary():
    """
    GET /api/v1/analytics/late-summary?year=2026&month=6
    Every active staff member with their late count for the month
    (staff with zero lates included). Sorted by late count desc.
    """
    try:
        year, month = _parse_year_month()
    except ValueError:
        return jsonify({"error": "Invalid year/month"}), 400

    month_start, month_end = _month_range(year, month)

    try:
        summary = _get_late_summary(year, month)
        return jsonify({
            "summary": summary,
            "year":    year,
            "month":   month,
            "start_date": str(month_start),
            "end_date":   str(month_end),
        }), 200

    except Exception as e:
        logger.error(f"Late summary error: {e}")
        return jsonify({"summary": [], "year": year, "month": month}), 200


# ── Monthly Late Summary Export (PDF / Excel) ──────────────────

@analytics_bp.route("/late-summary/export", methods=["GET"])
@admin_required
def late_summary_export():
    """
    GET /api/v1/analytics/late-summary/export?format=pdf|xlsx&year=&month=
    Downloads the monthly late summary as a PDF or Excel file.
    """
    fmt = request.args.get("format", "xlsx").lower()
    if fmt not in ("pdf", "xlsx"):
        return jsonify({"error": "format must be 'pdf' or 'xlsx'"}), 400

    try:
        year, month = _parse_year_month()
    except ValueError:
        return jsonify({"error": "Invalid year/month"}), 400

    try:
        summary = _get_late_summary(year, month)
    except Exception as e:
        logger.error(f"Late summary export query error: {e}")
        return jsonify({"error": "Failed to generate report"}), 500

    title    = f"Late Arrival Summary — {month_name[month]} {year}"
    filename = f"late_summary_{year}_{month:02d}.{fmt}"
    headers  = ["Employee ID", "Name", "Department",
                "Times Late", "Total Late (mins)", "Days Present"]
    data_rows = [
        [s["employee_id"], s["name"], s["department"] or "—",
         s["late_count"], s["total_late_minutes"], s["days_present"]]
        for s in summary
    ]

    try:
        if fmt == "xlsx":
            buf = _build_late_summary_xlsx(title, headers, data_rows)
            mimetype = ("application/vnd.openxmlformats-officedocument"
                        ".spreadsheetml.sheet")
        else:
            buf = _build_late_summary_pdf(title, headers, data_rows)
            mimetype = "application/pdf"
    except Exception as e:
        logger.error(f"Late summary export build error: {e}")
        return jsonify({"error": "Failed to generate report"}), 500

    return send_file(
        buf,
        mimetype=mimetype,
        as_attachment=True,
        download_name=filename,
    )


def _build_late_summary_xlsx(title: str, headers: list, rows: list) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Late Summary"

    # Title row
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=14)
    cell.alignment = Alignment(horizontal="center")

    # Header row
    header_fill = PatternFill("solid", fgColor="1E3A8A")
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = header_fill

    # Data rows
    for r, row in enumerate(rows, start=4):
        for col, val in enumerate(row, start=1):
            ws.cell(row=r, column=col, value=val)

    # Column widths
    widths = [14, 28, 20, 12, 18, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _build_late_summary_pdf(title: str, headers: list, rows: list) -> io.BytesIO:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import (SimpleDocTemplate, Table, TableStyle,
                                    Paragraph, Spacer)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        leftMargin=15 * mm, rightMargin=15 * mm,
        topMargin=15 * mm, bottomMargin=15 * mm,
        title=title,
    )

    styles = getSampleStyleSheet()
    elements = [
        Paragraph(title, styles["Title"]),
        Paragraph(f"Generated on {date.today().isoformat()}", styles["Normal"]),
        Spacer(1, 8),
    ]

    table_data = [headers] + [[str(v) for v in row] for row in rows]
    table = Table(table_data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1E3A8A")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.white),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 9),
        ("ALIGN",         (3, 1), (-1, -1), "CENTER"),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#CBD5E1")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.white, colors.HexColor("#F1F5F9")]),
        ("TOPPADDING",    (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(table)

    doc.build(elements)
    buf.seek(0)
    return buf


# ── Overtime Report ────────────────────────────────────────────

@analytics_bp.route("/overtime", methods=["GET"])
@admin_required
def overtime_report():
    """GET /api/v1/analytics/overtime?days=30"""
    days       = int(request.args.get("days", 30))
    end_date   = date.today()
    start_date = end_date - timedelta(days=days - 1)

    try:
        rows = db.session.execute(text("""
            SELECT
                e.employee_id,
                e.first_name || ' ' || e.last_name AS name,
                SUM(al.overtime_minutes)  AS total_overtime_minutes,
                SUM(al.working_minutes)   AS total_working_minutes,
                COUNT(*) AS days_worked
            FROM attendance_logs al
            JOIN employees e ON al.employee_id = e.id
            WHERE al.attendance_date BETWEEN :start AND :end
              AND al.overtime_minutes > 0
            GROUP BY e.employee_id, e.first_name, e.last_name
            ORDER BY total_overtime_minutes DESC
            LIMIT 20
        """), {"start": start_date, "end": end_date}).fetchall()

        report = []
        for row in rows:
            d = dict(row._mapping)
            report.append({
                "employee_id":           d["employee_id"],
                "name":                  d["name"],
                "total_overtime_minutes": int(d["total_overtime_minutes"] or 0),
                "total_working_minutes":  int(d["total_working_minutes"] or 0),
                "days_worked":            int(d["days_worked"] or 0),
            })

        return jsonify({
            "report":     report,
            "start_date": str(start_date),
            "end_date":   str(end_date),
        }), 200

    except Exception as e:
        logger.error(f"Overtime error: {e}")
        return jsonify({"report": []}), 200